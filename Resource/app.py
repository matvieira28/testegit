from openpyxl   import workbook, load_workbook


planilha = load_workbook("gerenciamento-de-banca.xlsx")
aba_ativa = planilha.active

for celula in aba_ativa["C"]:
    print(celula.value)



















# criar uma planilha (book)
# book = openpyxl.Workbook()

#como visualizar paginas
# print(book.sheetnames)

#como criar a propria página

# book.create_sheet("Frutas")

# frutas_page = book["Frutas"]

# frutas_page.a(["banana","5"])


#Salvar a planilha
# book.save("gerenciamento-de-banca-teste.xlsx")